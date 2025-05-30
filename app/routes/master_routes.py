from flask import Blueprint, request, jsonify
from openpyxl import Workbook
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from openpyxl.styles import PatternFill
from app.validation.master_validation import *
from app.dao.master_dao import *
from app.dao.quiz_dao import fetch_all_topics as quiz_all_topics
import bcrypt
import os

master_bp = Blueprint('master', __name__)


@master_bp.route('/topic_sample_excel', methods=['POST'])
def topic_sample_excel():
    try:
        # Optional: Validate the incoming JSON if required
        data = request.json
        if not validate_category_sample_excel(data):
            return jsonify({
                'success': False,
                'status': 1001,
                'message': 'Some fields are missing.',
                'response': None
            })

        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Topics"

        # Add headers
        ws.append(["Topic Name", "Description", "Categories"])
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40

        # Add example rows
        ws.append(["Math", "Basic mathematics", "Algebra,Geometry"])

        # Add additional blank rows for user input
        for _ in range(8):
            ws.append(["", "", ""])

        # Save the Excel file
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"topic_sample_{timestamp}.xlsx"
        save_path = os.path.join(os.getcwd(), "public", "downloads")
        os.makedirs(save_path, exist_ok=True)
        full_path = os.path.join(save_path, filename)
        wb.save(full_path)

        return jsonify({
            'success': True,
            'status': 200,
            'message': 'Excel generated successfully.',
            'response': {"filename": filename}
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Error generating Excel file.',
            'response': str(e)
        })


@master_bp.route('/question_sample_excel', methods=['POST'])
def question_sample_excel():
    try:
        data = request.json
        if not validate_category_sample_excel(data):
            return jsonify({
                'success': False,
                'status': 1001,
                'message': 'Some fields are missing.',
                'response': None
            })

        topics = quiz_all_topics(data)
        if len(topics) == 0:
            return jsonify({
                'success': False,
                'status': 500,
                'message': 'No topics found. Please add topics first.',
                'response': None
            })

        levels = fetch_levels(data)
        if len(levels) == 0:
            return jsonify({
                'success': False,
                'status': 500,
                'message': 'No levels found. Please add levels first.',
                'response': None
            })

        categories = get_all_categories(data)

        wb = Workbook()
        ws = wb.active
        ws.title = "Questions"

        # Define headers
        headers = [
            "Question", "Question Type", "Descriptive Answer",
            "Options", "Correct Answer", "Level",
            "Topic Name", "Category"
        ]
        ws.append(headers)

        # Set column widths
        column_widths = [20, 20, 20, 20, 20, 20, 25, 25]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Sample data
        ws.append([
            "What is the capital of France?", "MCQ", "",
            "Paris|London|Berlin", "Paris", levels[0]["name"],
            topics[0]["name"], ""
        ])
        ws.append([
            "Explain the process of photosynthesis.", "Descriptive",
            "Photosynthesis is the process...", "", "",
            levels[0]["name"], topics[0]["name"], ""
        ])

        # Create hidden sheet with lists
        hidden = wb.create_sheet(title="Lists")
        hidden.sheet_state = 'hidden'

        # Fill values
        topic_names = list(set([t['name'] for t in topics])) if topics else []
        category_names = list(set([c['categoryname'] for c in categories])) if categories else []
        level_names = list(set([l['name'] for l in levels])) if levels else []

        for idx, name in enumerate(topic_names, start=1):
            hidden.cell(row=idx, column=1, value=name)
        for idx, name in enumerate(category_names, start=1):
            hidden.cell(row=idx, column=2, value=name)
        for idx, name in enumerate(level_names, start=1):
            hidden.cell(row=idx, column=3, value=name)

        # Named ranges
        wb.create_named_range('TopicList', hidden, f"$A$1:$A${len(topic_names)}")
        wb.create_named_range('CategoryList', hidden, f"$B$1:$B${len(category_names)}")
        wb.create_named_range('LevelList', hidden, f"$C$1:$C${len(level_names)}")

        # Data Validations
        dv_qtype = DataValidation(type="list", formula1='"MCQ,Descriptive"', allow_blank=False)
        dv_level = DataValidation(type="list", formula1='=LevelList', allow_blank=False)
        dv_topic = DataValidation(type="list", formula1='=TopicList', allow_blank=False)
        dv_category = DataValidation(type="list", formula1='=CategoryList', allow_blank=False)

        # Add validations
        ws.add_data_validation(dv_qtype)
        ws.add_data_validation(dv_level)
        ws.add_data_validation(dv_topic)
        ws.add_data_validation(dv_category)

        for row in range(2, 12):  # Example: Apply to first 10 rows
            dv_qtype.add(ws[f"B{row}"])
            dv_level.add(ws[f"F{row}"])
            dv_topic.add(ws[f"G{row}"])
            dv_category.add(ws[f"H{row}"])

        # Save
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"question_sample_{timestamp}.xlsx"
        save_path = os.path.join(os.getcwd(), "public", "downloads")
        os.makedirs(save_path, exist_ok=True)
        full_path = os.path.join(save_path, filename)
        wb.save(full_path)

        return jsonify({
            'success': True,
            'status': 200,
            'message': 'Sample question Excel generated successfully.',
            'response': {"filename": filename}
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Error generating Excel file.',
            'response': str(e)
        })
    
@master_bp.route('/user_sample_excel', methods=['POST'])
def user_sample_excel():
    try:
        data = request.json
        if not validate_category_sample_excel(data):
            return jsonify({
                'success': False,
                'status': 1001,
                'message': 'Some fields are missing.',
                'response': None
            })

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"

        # Define headers
        headers = ["Name", "Email", "Username"]
        ws.append(headers)

        # Set column widths
        column_widths = [25, 35, 25]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Add sample data
        sample_users = [
            ["John Doe", "john.doe@example.com", "john.doe@example.com"],
            ["Jane Smith", "jane.smith@example.com", "jane.smith@example.com"]
        ]
        for row in sample_users:
            ws.append(row)

        # Align center
        # for row in ws.iter_rows(min_row=2, max_col=3, max_row=3):
        #     for cell in row:
        #         cell.alignment = Alignment(horizontal="left")

        # Save file
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"user_sample_{timestamp}.xlsx"
        save_path = os.path.join(os.getcwd(), "public", "downloads")
        os.makedirs(save_path, exist_ok=True)
        full_path = os.path.join(save_path, filename)
        wb.save(full_path)

        return jsonify({
            'success': True,
            'status': 200,
            'message': 'User sample Excel generated successfully.',
            'response': {"filename": filename}
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Error generating user Excel file.',
            'response': str(e)
        })
    

@master_bp.route('/upload_question_excel', methods=['POST'])
def upload_question_excel():
    try:
        if not validate_excel_upload_request(request, '.xlsx'):
            return jsonify({
                'success': False,
                'status': 1001,
                'message': 'Invalid request. Please check the uploaded file and required fields.',
                'response': None
            })

        clientid = request.form.get('clientid')
        file = request.files['file']
        wb = load_workbook(file)
        ws = wb.active

        # Fetch reference mappings
        topic_list = quiz_all_topics({'clientid': clientid})
        level_list = fetch_levels({'clientid': clientid})
        category_list = get_all_categories({'clientid': clientid})

        topics = {t['name']: t['topic_id'] for t in topic_list}
        levels = {l['name']: l['level_id'] for l in level_list}
        categories = {c['categoryname']: c['categoryid'] for c in category_list}
        print(f"Categories: {categories}")
        category_topic_map = {c['categoryname']: c['topicid'] for c in category_list}

        allowed_qtypes = ['MCQ', 'Descriptive']

        errors = []
        valid_data = []
        all_rows = []  # Will store (row_index, row_data_dict, errors_list)

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Skip empty rows
            print(f"Processing row {idx}: {row}")
            question, qtype, desc_ans, options, correct_ans, level_name, topic_name, category_name = row
            row_errors = []

            # Validate required fields
            if not question or not qtype or not level_name or not topic_name:
                row_errors.append("Question, Question Type, Level, and Topic are required.")

            # Validate dropdown values
            if qtype and qtype not in allowed_qtypes:
                row_errors.append("Invalid Question Type. Must be MCQ or Descriptive.")
            if level_name and level_name not in levels:
                row_errors.append(f"Invalid Level: '{level_name}'. Must be selected from dropdown.")
            if topic_name and topic_name not in topics:
                row_errors.append(f"Invalid Topic: '{topic_name}'. Must be selected from dropdown.")

            topic_id = topics.get(topic_name)
            category_id = categories.get(category_name) if category_name else None

            # if category_name:
            #     if category_name not in categories:
            #         row_errors.append(f"Invalid Category: '{category_name}'. Must be selected from dropdown or left blank.")
            #     elif category_topic_map.get(category_name) != topic_id:
            #         row_errors.append(f"Category '{category_name}' does not belong to selected topic '{topic_name}'.")

            if topic_name in topics:
                topic_id = topics.get(topic_name)
                topic_has_categories = any(cat['topicid'] == topic_id for cat in category_list)

                if topic_has_categories and not category_name:
                    row_errors.append(f"Topic '{topic_name}' has categories. Category is required.")

                if category_name:
                    if category_name not in categories:
                        row_errors.append(f"Invalid Category: '{category_name}'. Must be selected from dropdown.")
                    elif category_topic_map.get(category_name) != topic_id:
                        row_errors.append(f"Category '{category_name}' does not belong to selected topic '{topic_name}'.")


            # Additional logic checks
            if qtype == 'MCQ':
                if not options or not correct_ans:
                    row_errors.append("MCQ must have Options and Correct Answer.")
            elif qtype == 'Descriptive':
                if not desc_ans:
                    row_errors.append("Descriptive questions must have an answer.")

            row_data = {
                'clientid': int(clientid),
                'question': str(question).strip() if question else '',
                'question_type': 1 if qtype == 'MCQ' else 2 if qtype == 'Descriptive' else None,
                'descriptive_answer': str(desc_ans).strip() if desc_ans else '',
                'options': [opt.strip() for opt in str(options).split("|")] if options else [],
                'correct_answer': str(correct_ans).strip() if correct_ans else '',
                'level_id': levels.get(level_name),
                'topic_id': topic_id,
                'category_id': category_id
            }

            all_rows.append((idx, row_data, row_errors))

            if row_errors:
                errors.append({'row': idx, 'errors': row_errors})
            else:
                valid_data.append(row_data)

        total = len(all_rows)
        valid = len(valid_data)
        valid_percent = (valid / total) * 100 if total > 0 else 0

        # If all valid, insert directly
        if valid == total:
            for item in valid_data:
                insert_question(item)
            return jsonify({
                'success': True,
                'status': 200,
                'message': f"All {valid} questions inserted successfully.",
                'response': None
            })

        # Create error-highlighted Excel file with all rows + reason column
        error_filename = f"question_upload_errors_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        error_filepath = os.path.join(os.getcwd(), 'public', 'downloads')
        os.makedirs(error_filepath, exist_ok=True)
        full_error_path = os.path.join(error_filepath, error_filename)

        error_wb = Workbook()
        error_ws = error_wb.active
        error_ws.append([
            "Question", "Question Type", "Descriptive Answer",
            "Options", "Correct Answer", "Level",
            "Topic", "Category", "Reason"
        ])

        # Column widths for better readability
        col_widths = [30, 15, 30, 40, 20, 15, 25, 25, 50]
        for i, width in enumerate(col_widths, start=1):
            error_ws.column_dimensions[chr(64 + i)].width = width

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Write rows with errors highlighted
        for i, (row_num, data, errs) in enumerate(all_rows, start=2):
            row = [
                data['question'],
                'MCQ' if data['question_type'] == 1 else 'Descriptive' if data['question_type'] == 2 else '',
                data['descriptive_answer'],
                ", ".join(data['options']) if data['options'] else '',
                data['correct_answer'],
                next((k for k,v in levels.items() if v == data['level_id']), ''),
                next((k for k,v in topics.items() if v == data['topic_id']), ''),
                next((k for k,v in categories.items() if v == data['category_id']), '') if data['category_id'] else '',
                ", ".join(errs) if errs else ""
            ]
            error_ws.append(row)
            if errs:
                for col in range(1, 10):  # Highlight columns A to I
                    error_ws.cell(row=i, column=col).fill = red_fill

        error_wb.save(full_error_path)

        # Optionally log the error file somewhere
        insert_error_log({"type": "question", "filename": error_filename}, clientid)

        # Insert valid rows if >= 80%
        if valid_percent >= 80:
            for item in valid_data:
                insert_question(item)
            return jsonify({
                'success': True,
                'status': 200,
                'message': f"{valid} out of {total} questions inserted. Errors logged in file.",
                'response': {'error_file': error_filename}
            })
        else:
            return jsonify({
                'success': False,
                'status': 500,
                'message': "Less than 80% of rows are valid. No data inserted.",
                'response': {'error_file': error_filename}
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Error processing the Excel file.',
            'response': str(e)
        })






@master_bp.route('/upload_topics_excel', methods=['POST'])

def upload_topics_excel():
    try:
        # Get clientid from the form data
        if not validate_excel_upload_request(request, '.xlsx'):
            return jsonify({
                'success': False,
                'status': 1001,
                'message': 'Invalid request. Please check the uploaded file and required fields.',
                'response': None
            })

        clientid = request.form.get('clientid')
        file = request.files['file']

        wb = load_workbook(file)
        ws = wb.active

        errors = []
        valid_data = []
        all_rows = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            topic_name, description, categories = row
            row_errors = []

            if not topic_name or not str(topic_name).strip():
                row_errors.append("Topic Name is required.")
            if not description or not str(description).strip():
                row_errors.append("Description is required.")

            row_data = {
                'clientid': int(clientid),
                'topic_name': str(topic_name).strip() if topic_name else '',
                'description': str(description).strip() if description else '',
                'categories': [cat.strip() for cat in str(categories).split(",")] if categories else []
            }

            all_rows.append((idx, row_data, row_errors))

            if row_errors:
                errors.append({'row': idx, 'errors': row_errors})
            else:
                valid_data.append(row_data)

        total = len(all_rows)
        valid = len(valid_data)
        valid_percent = (valid / total) * 100 if total > 0 else 0

        # If all rows are valid
        if valid == total:
            for item in valid_data:
                result = insert_topic(item, clientid)
                if result is None:
                    return jsonify({
                        'success': False,
                        'status': 500,
                        'message': f"Error inserting topic: {item['topic_name']}",
                        'response': None
                    })

            return jsonify({
                'success': True,
                'status': 200,
                'message': f"All {valid} topics inserted successfully.",
                'response': None
            })

        # Create error-highlighted Excel
        error_filename = f"topic_upload_errors_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        error_filepath = os.path.join(os.getcwd(), 'public', 'downloads')
        os.makedirs(error_filepath, exist_ok=True)
        full_error_path = os.path.join(error_filepath, error_filename)

        error_wb = Workbook()
        error_ws = error_wb.active
        error_ws.append(["Topic Name", "Description", "Categories", "Reason"])

        # Set column widths
        error_ws.column_dimensions['A'].width = 30  # Topic Name
        error_ws.column_dimensions['B'].width = 50  # Description
        error_ws.column_dimensions['C'].width = 40  # Categories
        error_ws.column_dimensions['D'].width = 60  # Reason

        red_fill = PatternFill(start_color='FFC7CE',
                               end_color='FFC7CE', fill_type='solid')

        for i, (idx, data, errs) in enumerate(all_rows, start=2):
            row = [data['topic_name'], data['description'], ", ".join(
                data['categories']), ", ".join(errs) if errs else ""]
            error_ws.append(row)
            if errs:
                for col in range(1, 4):
                    error_ws.cell(row=i, column=col).fill = red_fill

        error_wb.save(full_error_path)

        # Log error file
        insert_error_log(
            {"type": "topic", "filename": error_filename}, clientid)

        if valid_percent >= 80:
            for item in valid_data:
                insert_topic(item)
            return jsonify({
                'success': True,
                'status': 200,
                'message': f"{valid} out of {total} topics inserted. Errors logged in file.",
                'response': {'error_file': error_filename}
            })
        else:
            return jsonify({
                'success': False,
                'status': 500,
                'message': f"Less than 80% valid rows. No data inserted.",
                'response': {'error_file': error_filename}
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Error processing file.',
            'response': str(e)
        })
    
@master_bp.route('/upload_users', methods=['POST'])
def upload_users():
    try:
        if not validate_excel_upload_request(request, '.xlsx'):
            return jsonify({
                'success': False,
                'status': 1001,
                'message': 'Invalid request. Please check the uploaded file and required fields.',
                'response': None
            })

        clientid = request.form.get('clientid')
        file = request.files['file']

        wb = load_workbook(file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        expected_headers = ["Name", "Email", "Username"]
        if headers != expected_headers:
            return jsonify({
                'success': False,
                'status': 400,
                'message': f'Invalid Excel format. Expected headers: {expected_headers}.',
                'response': None
            })

        errors = []
        valid_data = []
        all_rows = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name, email, username = row
            row_errors = []

            if not name or not str(name).strip():
                row_errors.append("Name is required.")
            if not email or not str(email).strip():
                row_errors.append("Email is required.")
            if not username or not str(username).strip():
                row_errors.append("Username is required.")

            row_data = {
                'clientid': int(clientid),
                'name': str(name).strip() if name else '',
                'email': str(email).strip() if email else '',
                'username': str(username).strip() if username else '',
                'password': bcrypt.hashpw('password'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),
                'usertype': 3
            }

            all_rows.append((idx, row_data, row_errors))

            if row_errors:
                errors.append({'row': idx, 'errors': row_errors})
            else:
                valid_data.append(row_data)

        total = len(all_rows)
        valid = len(valid_data)
        valid_percent = (valid / total) * 100 if total > 0 else 0
        print(f"Total rows: {total}, Valid rows: {valid}, Valid percentage: {valid_percent:.2f}%")

        # If all rows are valid
        if valid == total:
            insert_values = []
            for data in valid_data:
                insert_values.append((
                    clientid, data['email'], data['name'], data['username'], data['password'],
                    None, data['usertype']
                ))

            result = insert_users(insert_values)
            if result:
                return jsonify({
                    'success': True,
                    'status': 200,
                    'message': f"All {valid} users inserted successfully.",
                    'response': None
                })
            else:
                return jsonify({
                    'success': False,
                    'status': 500,
                    'message': 'Error inserting users.',
                    'response': None
                })

        # Generate error Excel file
        error_filename = f"user_upload_errors_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        error_filepath = os.path.join(os.getcwd(), 'public', 'downloads')
        os.makedirs(error_filepath, exist_ok=True)
        full_error_path = os.path.join(error_filepath, error_filename)

        error_wb = Workbook()
        error_ws = error_wb.active
        error_ws.append(["Name", "Email", "Username", "Reason"])

        # Set column widths
        error_ws.column_dimensions['A'].width = 30
        error_ws.column_dimensions['B'].width = 30
        error_ws.column_dimensions['C'].width = 30
        error_ws.column_dimensions['D'].width = 60

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        for i, (idx, data, errs) in enumerate(all_rows, start=2):
            row = [data['name'], data['email'], data['username'], ", ".join(errs)]
            error_ws.append(row)
            if errs:
                for col in range(1, 4):
                    error_ws.cell(row=i, column=col).fill = red_fill

        error_wb.save(full_error_path)

        # Log error file
        insert_error_log({"type": "user", "filename": error_filename}, clientid)

        # Insert if valid ≥ 80%
        if valid_percent >= 80:
            insert_values = []
            for data in valid_data:
                insert_values.append((
                    clientid, data['email'], data['name'], data['username'], data['password'],
                    None, data['usertype']
                ))

            insert_users(insert_values)

            return jsonify({
                'success': True,
                'status': 200,
                'message': f"{valid} out of {total} users inserted. Errors logged in file.",
                'response': {'error_file': error_filename}
            })

        else:
            return jsonify({
                'success': False,
                'status': 500,
                'message': f"Less than 80% valid rows. No data inserted.",
                'response': {'error_file': error_filename}
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Upload failed.',
            'response': str(e)
        })



@master_bp.route('/topics', methods=['POST'])
def topics():
    try:
        data = request.json
        if not validate_topic_data(data):
            return jsonify({
                'success': False,
                'status': 400,
                'message': 'Every field is required to start.',
                'response': None
            })

        clientid = data.get('clientid')

        topics = fetch_all_topics(data)
        if topics is None:
            return jsonify({
                'success': False,
                'status': 500,
                'message': 'Couldn’t load topics right now. Please try again soon.',
                'response': None
            })
        

        # Fetch categories grouped by topicid
        category_rows = get_categories(clientid)

        # Build topicid -> categories mapping
        topic_categories_map = {
            row['topicid']: row['categories'] for row in category_rows}

        # Merge categories into each topic
        for topic in topics:
            topic_id = topic.get('topic_id')
            topic['categories'] = topic_categories_map.get(topic_id, '')

        return jsonify({
            'success': True,
            'status': 200,
            'message': '',
            'response': topics
        })

    except Exception as e:
        print(f"Error in /topics: {e}")
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Something went wrong. Please try again later.',
            'response': None
        })


@master_bp.route('/update_topic', methods=['POST'])
def update_topic():
    try:
        data = request.json
        if not validate_update_topic_data(data):
            return jsonify({
                'success': False,
                'status': 400,
                'message': 'Every field is required to update.',
                'response': None
            })

        result = update_topic_data(data)
        if not result:
            return jsonify({
                'success': False,
                'status': 500,
                'message': 'Couldn’t update image right now. Please try again soon.',
                'response': None
            })

        return jsonify({
            'success': True,
            'status': 200,
            'message': 'Image updated successfully.',
            'response': None
        })

    except Exception as e:
        print(f"Error in /topics: {e}")
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Something went wrong. Please try again later.',
            'response': None
        })


@master_bp.route('/topic_details', methods=['POST'])
def topic_details():
    try:
        data = request.json
        if not validate_topic_details(data):
            return jsonify({
                'success': False,
                'status': 400,
                'message': 'Every field is required to update.',
                'response': None
            })

        result = get_topic_details(data)
        if result is None:
            return jsonify({
                'success': False,
                'status': 500,
                'message': 'Couldn’t fetch. Please try again soon.',
                'response': None
            })

        return jsonify({
            'success': True,
            'status': 200,
            'message': 'Fetched successfully.',
            'response': result
        })

    except Exception as e:
        print(f"Error in /topics: {e}")
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Something went wrong. Please try again later.',
            'response': None
        })


@master_bp.route('/questions', methods=['POST'])
def questions():
    try:
        data = request.json
        if not validate_question_data(data):
            return jsonify({
                'success': False,
                'status': 400,
                'message': 'Every field is required to start.',
                'response': None
            })

        questions = fetch_all_questions(data)
        if questions is None:
            return jsonify({
                'success': False,
                'status': 500,
                'message': 'Couldn’t load questions right now. Please try again soon.',
                'response': None
            })

        return jsonify({
            'success': True,
            'status': 200,
            'message': '',
            'response': questions
        })

    except Exception as e:
        print(f"Error in /questions: {e}")
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Something went wrong. Please try again later.',
            'response': None
        })
    
@master_bp.route('/users', methods=['POST'])
def users():
    try:
        data = request.json
        if not validate_question_data(data):
            return jsonify({
                'success': False,
                'status': 400,
                'message': 'Every field is required to start.',
                'response': None
            })

        users = fetch_all_users(data)
        if users is None:
            return jsonify({
                'success': False,
                'status': 500,
                'message': 'Couldn’t load users right now. Please try again soon.',
                'response': None
            })

        return jsonify({
            'success': True,
            'status': 200,
            'message': '',
            'response': users
        })

    except Exception as e:
        print(f"Error in /users: {e}")
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Something went wrong. Please try again later.',
            'response': None
        })
    
@master_bp.route('/error_logs', methods=['POST'])
def error_logs():
    try:
        data = request.json
        if not validate_question_data(data):
            return jsonify({
                'success': False,
                'status': 400,
                'message': 'Every field is required to start.',
                'response': None
            })

        logs = fetch_all_error_logs(data)
        if logs is None:
            return jsonify({
                'success': False,
                'status': 500,
                'message': 'Couldn’t load error logs right now. Please try again soon.',
                'response': None
            })

        return jsonify({
            'success': True,
            'status': 200,
            'message': '',
            'response': logs
        })

    except Exception as e:
        print(f"Error in /error_logs: {e}")
        return jsonify({
            'success': False,
            'status': 500,
            'message': 'Something went wrong. Please try again later.',
            'response': None
        })
